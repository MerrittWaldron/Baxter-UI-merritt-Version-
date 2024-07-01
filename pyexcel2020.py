'''
Copyright (c) 2020 Paul Marichal

permission is hereby granted, free of charge, to any person obtaining a copy of
this software and associated documentation files (the "Software"), to deal in
the Software without restriction, including without limitation the rights to
use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies
of the Software, and to permit persons to whom the Software is furnished to do
so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
'''


#  pylint: disable=W0312, C0103

import os
import sys
import platform
import time
from collections import Counter
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo


InputDirectoryPath = './input'
OutputDirectoryPath = './output'

#  make sure file exist before we start


def check_if_mergeddata_exists(OutputDirectoryDataPathFN):
	newwb = Workbook()
	if os.path.exists(OutputDirectoryDataPathFN):
		print('Excel output exists')
	else:
		newwb.save(filename=OutputDirectoryDataPathFN)

# supports older Version of pandas. This method no longer used with new pandas version.

def append_df_to_excel1(filename, df, sheet_name='Sheet1', **to_excel_kwargs):
	"""
	Append a DataFrame [df] to existing Excel file [filename]
	into [sheet_name] Sheet.
	If [filename] doesn't exist, then this function will create it.

	Parameters:
		filename : File path or existing ExcelWriter
		(Example: '/path/to/file.xlsx')
		df : dataframe to save to workbook
		sheet_name : Name of sheet which will contain DataFrame.(default: 'Sheet1')
		to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
	Returns: None
	"""
	#  ignore [engine] parameter if it was passed
	if 'engine' in to_excel_kwargs:
		to_excel_kwargs.pop('engine')

	writer = pd.ExcelWriter(filename, engine='openpyxl')

	try:
		#  try to open an existing workbook
		writer.book = load_workbook(filename)
		#  see if sheet already exists
		if sheet_name in writer.book.sheetnames:
			#  read to existing sheet from Excel file
			dfold = pd.read_excel(open(filename, 'rb'), sheet_name, header=0)
			#  concatenate old and new dataframe
			df = pd.concat([df, dfold])
			#  delete original sheet we can create a new one with appended data
			bname = writer.book[sheet_name]
			writer.book.remove(bname)
			#  make sure there are no duplicate keys
			if 'Key' in sheet_name:
				df.drop_duplicates(subset='Batch', keep='first', inplace=True)
	except FileNotFoundError:
		#  file does not exist yet, we will create it
		pass

	#  write out the new sheet

	df.to_excel(writer, sheet_name, startrow=0, **to_excel_kwargs)
	#  save the workbook
	writer.save()
	#  now we want to change excel workbook sheet data to table
	mywb = openpyxl.load_workbook(filename)
	#  make sure the data is in Excel table format
	tab = openpyxl.worksheet.table.Table(displayName=sheet_name, ref=f'A1:{chr(len(df.columns)+64)}{len(df)+1}')
	#  print('EXCEL TABLE', chr(len(df.columns)+64), len(df)+1)
	style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=TableStyleInfo)
	tab.tableStyleInfo = style
	#  write out the new table
	mywb[sheet_name].add_table(tab)
	mywb.save(filename)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, startcol=None,
    truncate_sheet=False, resizeColumns=True, na_rep = 'NA', **to_excel_kwargs):
    """

    NEW Method
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file

      resizeColumns: default = True . It resize all columns based on cell content width
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
      na_rep: default = 'NA'. If, instead of NaN, you want blank cells, just edit as follows: na_rep=''


    Returns: None

    *******************

    CONTRIBUTION:
    Current helper function generated by [Baggio]: https://stackoverflow.com/users/14302009/baggio?tab=profile
    Contributions to the current helper function: https://stackoverflow.com/users/4046632/buran?tab=profile
    Original helper function: (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)


    Features of the new helper function:
    1) Now it works with python 3.9 and latest versions of pandas and openpxl
    ---> Fixed the error: "zipfile.BadZipFile: File is not a zip file".
    2) Now It resize all columns based on cell content width AND all variables will be visible (SEE "resizeColumns")
    3) You can handle NaN,  if you want that NaN are displayed as NaN or as empty cells (SEE "na_rep")
    4) Added "startcol", you can decide to start to write from specific column, oterwise will start from col = 0

    *******************



    """
    from openpyxl import load_workbook
    #from string import ascii_uppercase
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    try:
        f = open(filename)
        # Do something with the file
    except IOError:
        # print("File not accessible")
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename)

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace')


    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    # try:
    #     # try to open an existing workbook
    #     writer.book = load_workbook(filename)

    #     # get the last row in the existing Excel sheet
    #     # if it was not specified explicitly
    #     if startrow is None and sheet_name in writer.book.sheetnames:
    #         startrow = writer.book[sheet_name].max_row

    #     # truncate sheet
    #     if truncate_sheet and sheet_name in writer.book.sheetnames:
    #         # index of [sheet_name] sheet
    #         idx = writer.book.sheetnames.index(sheet_name)
    #         # remove [sheet_name]
    #         writer.book.remove(writer.book.worksheets[idx])
    #         # create an empty sheet [sheet_name] using old index
    #         writer.book.create_sheet(sheet_name, idx)

    #     # copy existing sheets
    #     writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    # except FileNotFoundError:
    #     # file does not exist yet, we will create it
    #     pass
    # if the sheet is not new then flag it for use later
    not_new_sheet = False
    for sname in writer.sheets:
        if sheet_name in sname:
            not_new_sheet = True

    if startrow is None:
        # startrow = -1
        startrow = 0

    if startcol is None:
        startcol = 0

    # write out the new sheet without header if the sheet is not new.
    if not_new_sheet:
        df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, na_rep=na_rep, header = True, **to_excel_kwargs)
    else:
        df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, na_rep=na_rep, header = True, **to_excel_kwargs)

    if resizeColumns:

        ws = writer.book[sheet_name]

        def auto_format_cell_width(ws):
            for letter in range(1,ws.max_column):
                maximum_value = 0
                for cell in ws[get_column_letter(letter)]:
                    val_to_check = len(str(cell.value))
                    if val_to_check > maximum_value:
                        maximum_value = val_to_check
                ws.column_dimensions[get_column_letter(letter)].width = maximum_value + 2

        auto_format_cell_width(ws)

    # save the workbook
    writer.save()


''' this method checks to see if the filename tracker file exist,
	creates it if it doesn't or returns a list of only the new files to be processed
	It create a new file or appends to existing '''


def find_new_filenames_to_upload(all_filenames_in_dir, txtfilename):
	if os.path.exists(txtfilename):
		with open(txtfilename) as f:
			file_content = f.readlines()
			converted_list = []
			for element in file_content:
				converted_list.append(element.strip())
			file_content = converted_list
		list1 = file_content
		list2 = all_filenames_in_dir
		C1 = Counter(list1)
		C2 = Counter(list2)
		new_filenames = list((C2 - C1).elements())
		return new_filenames
	else:
		f1 = open(txtfilename, 'w')
		f1.close()
		return all_filenames_in_dir


''' this method checks to see if the filename tracker file exist,
	creates it if it doesn't or returns a list of only the new files to be processed
	It create a new file or appends to existing '''


def write_FN_file(filetowite, txtfilename):
	if os.path.exists(txtfilename):
		f2 = open(txtfilename, 'a')
		print(" ADDING", filetowite)
		filetowite = filetowite + '\n'
		f2.write(filetowite)
		f2.close()
		return 1


''' this method processes Brew files and adds them to output file '''


def merge_excel_brew_files(InputBrewLogPath, OutputDirectoryDataPathFN):
	try:
		# warnings.filterwarnings("error", category=UserWarning)
		file_cnt = 0
		excel_names = []
		# warnings.filterwarnings("error", category=Warning)
		#  r=root, d=directories, f = files
		for r, d, f in os.walk(InputBrewLogPath):
			for file in f:
				if '.xlsx' in file:
					excel_names.append(os.path.join(r, file))
		if len(excel_names) == 0:
			print('\nNo files available to process')
			return 1

		# new_filenames = []
		txtfilename = OutputDirectoryPath + "/brewFN.txt"
		# new_filenames = find_new_filenames_to_upload(excel_names, txtfilename)
		# if len(new_filenames) == 0:
		# 	print('\nNo new files available to process')
		# 	return 1
		# excel_names = new_filenames
		#  read them into list
		excels = [pd.ExcelFile(name) for name in excel_names]


		#  print out filenames
		print("\n")
		# file_cnt = 0
		#  turn them into dataframes
		frames = []
		keydata = []
		batch_cnt = 0
		for x in excels:
			print("File processed", excel_names[file_cnt])
			file_cnt = file_cnt + 1
			for sname in x.sheet_names:
				# don't process the sheets that are not weekly brews
				if "Template" not in sname and "Lookup" not in sname and "List" not in sname \
					and "Malts" not in sname and "BH01" not in sname and "BH02" not in sname \
					and "Year" not in sname and "Print" not in sname:
					#  initial read to get batch and Date
					#  now read only data we are interested in
					print('Processing sheet', sname)
					# read in the file with selected columns only
					df = pd.read_excel(x, sname, header=None, nrows=5, usecols=[4, 5, 6, 7, 8, 9, 10])
					#  get the batch and date from original DF
					batch = df.iat[0, 0]
					mydate = df.iat[0, 5]
					print("    ", batch, mydate, " - processing")
					#  read again but restrict columns and rows that we will include
					#  skip the first 5 rows, then read the next 45 rows and only specific columns
					df = pd.read_excel(x, sname, header=[0], skiprows=5, nrows=38, usecols=[6, 7, 8, 9, 10])
					#  The row names are being read that will be used as headers
					dfns = pd.read_excel(x, sname, header=[0], skiprows=5, nrows=38, usecols=[6])
					# drop  all NaN values
					dfns.dropna(inplace=True)
					loop = 0
					# we are going to work on the Measurement column
					columnSeriesObjMeasure = dfns['Measurement']
					#  loop through row list and look for non-string value
					for i in columnSeriesObjMeasure:
						if isinstance(columnSeriesObjMeasure.values[loop], str):
							loop = loop + 1
							continue
						else:
							#  we need to stop processing files since data is not valid
							print('Non string value in Brew File', excel_names[file_cnt - 1], '--- Non String Value ---')
							print('Probably a non text string under Measurement column')
							print('Look for number value', columnSeriesObjMeasure.values[loop])
							print(columnSeriesObjMeasure)
							return 0
					#  reset headers to be that of names that are in the Measurement column
					df.set_index('Measurement', inplace=True)
					#  now make all the row names of column Measurement as new table headers
					dfn = df.transpose()
					#  insert new columns we are creating
					dfn.insert(0, "Brand", batch[0:3], True)
					dfn.insert(1, "Batch", batch, True)
					dfn.insert(2, "Date", mydate, True)
					#check to make sure the first stirke temp is not empty since this will create an empty dataframe
					value = dfn.iat[0, 4]
					debugshape = dfn.shape[1]
					# read in the last column to make sure the frame is correct by checking the column name. It should always be the same
					debugvalue = dfn.columns.values[40]
					# If the value is not correct, stop processing the frame because the Excel sheet if not formated correctly
					if debugvalue != "Original Gravity (Plato)":
						print("########################################################Frame does not have correct number of rows", sname, x)
						return 0
					# I drop columns that are NaN which usually means no batch was brewed
					# column 1 should always be populated but sometimes BH02 doesn't measure it so make it 999 so we don't drop the column
					isNaN = np.isnan(value)
					if isNaN :
						print('----------------------------------Strike Temp not populated')
						# if the value is NaN then assign it 999 to indicate no entry was done by brewer
						dfn.iat[0, 4] = 999
					dfnheaders = list(dfn.columns.values)
					oldframefmt = str(dfnheaders[5])
					# this code is here to make sure we convert all sheets to a consistent format
					# also make sure that the naming convention is the same for all files to be concatenated.
					
					#new code to fix column name changes. Make sure dataframe names are all the same for merging
					dfn.columns.values[5] = 'Blank 1'
					dfn.columns.values[6] = 'Strike water vol (gal)'
					dfn.columns.values[7] = 'Target strike vol (gal)'
					dfn.columns.values[17] = 'Sparge water vol (gal)'
					dfn.columns.values[18] = 'Target sparge (gal)'
					dfn.columns.values[19] = 'Kettle Temp Stop'
					dfn.columns.values[22] = 'Kettle Gravity Pre (Plato)'
					dfn.columns.values[33] = 'Kettle Gravity (Plato)'
					dfn.columns.values[35] = 'Cool pool Temperature'
					dfn.columns.values[36] = 'Castout Temperature'
					dfn.columns.values[38] = 'Castout vol (gal)'
					# 	#print("Updating to new format")

					# if "HLT" in oldframefmt:
					# 	dfn.columns.values[5] = 'Estimated Strike Temp'
					# 	dfn.columns.values[6] = 'HLT volume (Mash)'
					# 	dfn.columns.values[7] = 'Strike water vol (gal)'
					# 	dfn['HLT volume (Mash)'] = dfn['Estimated Strike Temp']
					# 	dfn['Estimated Strike Temp'] = 0
					# 	dfn['Target strike vol (gal)'] = 0
					# 	dfn.columns.values[18] = 'Target sparge (gal)'
					# 	dfn.columns.values[19] = 'Kettle Temp Stop'
					# 	dfn.columns.values[22] = 'Kettle Gravity Pre (Plato)'
					# 	dfn.columns.values[33] = 'Kettle Gravity (Plato)'
					# 	dfn.columns.values[35] = 'Cool pool Temperature'
					# 	dfn.columns.values[38] = 'Castout vol (gal)'
					# 	dfn['Kettle Temp Stop'] = dfn['Target sparge (gal)']
					# 	dfn['Target sparge (gal)'] = 0
					# 	#print("Updating to new format")
					# else:
					# 	# make sure all column headers are correct in case of typos
					# 	dfn.columns.values[5] = 'Estimated Strike Temp'
					# 	dfn.columns.values[6] = 'Strike water vol (gal)'
					# 	dfn.columns.values[7] = 'Target strike vol (gal)'
					# 	#dfn.columns.values[17] = 'Sparge water vol (gal)'
					# 	dfn.columns.values[18] = 'Target sparge (gal)'
					# 	dfn.columns.values[19] = 'Kettle Temp Stop'
					# 	dfn.columns.values[22] = 'Kettle Gravity Pre (Plato)'
					# 	dfn.columns.values[33] = 'Kettle Gravity (Plato)'
					# 	dfn.columns.values[35] = 'Cool pool Temperature'
					# 	dfn.columns.values[38] = 'Castout vol (gal)'
					# 	# print("Detected new format")


					# drop all the columns that don't have a Strike temp entry which identifies the number of turns
					dfn.dropna(axis=0, subset=['Strike Temp'], inplace=True)
					# print('Pre',dfn.shape)
					#  create a dataframe containing unique batch names
					#  this will be stored in a separate sheet and used as primary keys
					dashlst = [[batch[0:3], batch, mydate]]
					# create a dataframe with this data
					dfdash = pd.DataFrame(dashlst, columns=['Brand', 'Batch', 'Date'])
					# copy only the columns that have valid data in
					#  make sure we only get the rows that we want and have valid data
					dfn = dfn.iloc[:, [0, 1, 2, 3, 4, 6, 7, 10, 11, 13, 14, 16, 17, 18, 19, 20, 21, 22, 23, 31, 32, 33, 34, 35, 36, 38, 39, 40]].copy()
					#dfn.fillna(0, inplace=True)
					getdfnshape = dfn.shape[1]
					# when done processing, make sure there's exactly 28 columns
					if getdfnshape != 28:
						print('+++++++++++++++++++++++++++++++++Table not formated properly', dfn.shape, sname)
						return 0
					batch_cnt = batch_cnt + 1
					#  check to make sure all the headers names match to the previous dataframe
					# if not report an error
					if len(frames) > 1:
						dfndiff = dfn.columns.difference(frames[-1].columns)
						if dfndiff.empty is False:
							print('Frame column headers do not match=========================================================', dfndiff)
					#  create a list of appended dataframes
					frames.append(dfn)
					keydata.append(dfdash)
			write_FN_file(excel_names[file_cnt - 1], txtfilename)
		print("Processed ", file_cnt, " Brew files and ", batch_cnt, " Batches")
		#  concatenate all the frames
		combined = pd.concat(frames, sort=False)
		combined = combined.iloc[:, 0:28].copy()
		# create the Key sheet for Excel
		combined1 = pd.concat(keydata, sort=False)
		combined1.drop_duplicates(subset='Batch', keep='first', inplace=True)
		#  write it out to Excel file
		print("Appending brew sheets to Excel file")
		append_df_to_excel(OutputDirectoryDataPathFN, combined, sheet_name='Brew', index=False)
		print("Appending key sheets to Excel file")
		append_df_to_excel(OutputDirectoryDataPathFN, combined1, sheet_name='Key', index=False)
		print("Updating brewFN.txt file")
		return 1
	except Exception as e:
		print("Entry in Brew file", excel_names[file_cnt - 1], sname, "is not valid")
		print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)
		# print("Exception", e)
		return 0


def print_menu():
	""" print initial menu for user"""
	print(30 * "-" + "MENU" + 30 * "-")
	print("1   - Merge Filter Logs")
	print("2   - Merge Tank Logs")
	print("3   - Merge Brew Logs")
	print("4   - Quit")
	print(67 * "-")


if __name__ == "__main__":
	# programinput_path = sys.argv[1]
	# jirainput_path = sys.argv[2]
	# production = sys.argv[3]
	InputFilterLogFN = '/Filter'
	InputBrewLogFN = '/Brew'
	InputTankLogFN = '/Tank'
	OutputmergeddataFN = '/MergedData'

	InputFilterterLogPath = InputDirectoryPath + InputFilterLogFN + '/'
	InputBrewLogPath = InputDirectoryPath + InputBrewLogFN + '/'
	InputTankLogPath = InputDirectoryPath + InputTankLogFN + '/'
	OutputDirectoryDataPathFN = OutputDirectoryPath + OutputmergeddataFN + '.xlsx'

	print("Python Version from is " + platform.python_version())
	print("System Version is " + platform.platform())

	print("Software Version is V6.5.1")

	localtime = time.asctime(time.localtime(time.time()))
	print("Local current time :", localtime)
	check_if_mergeddata_exists(OutputDirectoryDataPathFN)

	while True:
		print_menu()
		try:
			choice = int(input("Enter your choice [1-4]: "))
		except ValueError:
			print("Not a valid number selection")
			#  better try again... Return to the start of the loop
			continue
		if choice < 1 or choice > 4:
			print("Selection not in range")
			continue
		if choice == 1:
			merge_excel_filter_files(InputFilterterLogPath, OutputDirectoryDataPathFN)
		elif choice == 2:
			merge_excel_tank_files(InputTankLogPath, OutputDirectoryDataPathFN)
		elif choice == 3:
			merge_excel_brew_files(InputBrewLogPath, OutputDirectoryDataPathFN)
		elif choice == 4:
			break
